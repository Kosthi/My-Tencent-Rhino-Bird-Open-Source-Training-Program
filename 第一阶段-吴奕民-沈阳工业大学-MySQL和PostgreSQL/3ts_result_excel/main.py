import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter


def process_file(file_path):
    data = {}
    with open(file_path, 'r') as file:
        for line in file:
            if ':' in line:
                key, value = line.strip().split(':', 1)
                data[key.strip()] = value.strip()
    return data


def get_database_and_isolation(folder, file_name):
    parts = file_name.split('_')
    database = folder.replace('_', ' ')
    isolation = parts[0].replace('-', ' ').title()
    return database, isolation


all_data = []
folders = ['MySQL_8.0.39', 'PostgreSQL_12.20']

for folder in folders:
    summary_folder = os.path.join(folder, 'result_summary')
    if os.path.exists(summary_folder):
        for filename in os.listdir(summary_folder):
            if filename.endswith('total-result.txt'):
                file_path = os.path.join(summary_folder, filename)
                data = process_file(file_path)
                database, isolation = get_database_and_isolation(folder, filename)
                for test_case, handling in data.items():
                    all_data.append({
                        'Database': database,
                        'Isolation Level': isolation,
                        'Test Case': test_case,
                        'Handling': 'Pass' if handling == 'Avoid' else handling
                    })

df = pd.DataFrame(all_data)

pivot_df = df.pivot_table(
    values='Handling',
    index='Test Case',
    columns=['Database', 'Isolation Level'],
    aggfunc='first'
)

output_filename = 'transaction_summary.xlsx'
pivot_df.to_excel(output_filename)

# 使用 openpyxl 调整单元格大小
wb = load_workbook(output_filename)
ws = wb.active

# 设置列宽
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column_letter].width = adjusted_width

# 设置行高
for row in ws.rows:
    ws.row_dimensions[row[0].row].height = 20

# 设置所有单元格居中对齐
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 定义颜色映射
color_map = {
    'Pass': 'CCFFCC',  # 浅绿色
    'Anomaly': 'FFCCCC',  # 浅红色
    'Rollback': 'FFFFCC',  # 浅黄色
}

# 定义字体
times_new_roman = Font(name='Times New Roman', size=11)

# 定义边框样式
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# 遍历所有数据单元格并设置字体、边框和颜色
for row in ws.iter_rows(min_row=4, min_col=2):  # 跳过标题行和索引列
    for cell in row:
        cell.border = thin_border
        cell.font = times_new_roman
        if cell.value in color_map:
            cell.fill = PatternFill(start_color=color_map[cell.value], end_color=color_map[cell.value],
                                    fill_type='solid')

# 保存调整后的文件
wb.save(output_filename)

print(f"数据已保存到 '{output_filename}'")
